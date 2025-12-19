"""
Quality Assurance Export for Extraction Verification
Exports case data in a format comparable to the Excel standard for QA review.

Usage:
    python -m pipeline.qa_export --output qa_review.csv --limit 100
    python -m pipeline.qa_export --case-ids 1,2,3,4,5
    python -m pipeline.qa_export --output qa_review.xlsx --format excel
"""

import argparse
import csv
import logging
import json
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any

from sqlalchemy import create_engine, text

from .config import Config

logger = logging.getLogger(__name__)


class QAExporter:
    """
    Export extracted case data for quality assurance verification.
    Produces output comparable to Excel standard for manual review.
    """
    
    def __init__(self, db_url: Optional[str] = None):
        self.db_url = db_url or Config.get_database_url()
        self.engine = create_engine(self.db_url)
    
    def export_cases(
        self,
        output_path: str,
        case_ids: Optional[List[int]] = None,
        limit: Optional[int] = None,
        format: str = 'csv'
    ) -> int:
        """
        Export cases to CSV or Excel format.
        
        Args:
            output_path: Output file path
            case_ids: Specific case IDs to export (optional)
            limit: Maximum number of cases (optional)
            format: 'csv' or 'excel'
            
        Returns:
            Number of cases exported
        """
        cases = self._fetch_cases(case_ids, limit)
        
        if format == 'excel':
            return self._write_excel(cases, output_path)
        else:
            return self._write_csv(cases, output_path)
    
    def _fetch_cases(
        self, 
        case_ids: Optional[List[int]] = None,
        limit: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """Fetch case data with all related entities."""
        
        # Build WHERE clause
        where_clause = ""
        params = {}
        
        if case_ids:
            where_clause = "WHERE c.case_id = ANY(:case_ids)"
            params['case_ids'] = case_ids
        
        limit_clause = f"LIMIT {limit}" if limit else ""
        
        query = text(f"""
            SELECT 
                c.case_id,
                c.docket_number,
                c.title,
                c.court_level,
                c.court,
                c.county,
                c.district,
                c.source_docket_number,
                c.appeal_outcome,
                c.winner_legal_role AS case_winner_legal,
                c.winner_personal_role AS case_winner_personal,
                c.case_type,
                c.decision_year,
                c.decision_month,
                c.publication_status,
                c.issue_count,
                c.source_file,
                c.extraction_timestamp,
                
                -- Parties (aggregated)
                (SELECT json_agg(json_build_object(
                    'name', p.name,
                    'role', p.legal_role,
                    'party_type', p.party_type,
                    'personal_role', p.personal_role
                ))
                FROM parties p WHERE p.case_id = c.case_id) AS parties,
                
                -- Judges (aggregated)
                (SELECT json_agg(json_build_object(
                    'name', j.name,
                    'role', cj.role
                ))
                FROM case_judges cj
                JOIN judges j ON cj.judge_id = j.judge_id
                WHERE cj.case_id = c.case_id) AS judges,
                
                -- Issues (aggregated)
                (SELECT json_agg(json_build_object(
                    'issue_id', id.issue_id,
                    'summary', id.issue_summary,
                    'outcome', id.issue_outcome,
                    'winner_legal', id.winner_legal_role,
                    'winner_personal', id.winner_personal_role,
                    'decision_summary', id.decision_summary,
                    'taxonomy', lt.name,
                    'rcw_reference', id.rcw_reference,
                    'keywords', id.keywords,
                    'confidence', id.confidence_score
                ) ORDER BY id.issue_id)
                FROM issues_decisions id
                LEFT JOIN legal_taxonomy lt ON id.taxonomy_id = lt.taxonomy_id
                WHERE id.case_id = c.case_id) AS issues,
                
                -- RCW count from junction table
                (SELECT COUNT(DISTINCT ir.rcw_id) 
                 FROM issue_rcw ir 
                 JOIN issues_decisions id ON ir.issue_id = id.issue_id 
                 WHERE id.case_id = c.case_id) AS rcw_count,
                
                -- Citation count
                (SELECT COUNT(*) FROM citation_edges ce WHERE ce.source_case_id = c.case_id) AS citation_count,
                
                -- Statute count
                (SELECT COUNT(*) FROM statute_citations sc WHERE sc.case_id = c.case_id) AS statute_count
                
            FROM cases c
            {where_clause}
            ORDER BY c.case_id DESC
            {limit_clause}
        """)
        
        with self.engine.connect() as conn:
            result = conn.execute(query, params)
            rows = result.fetchall()
            columns = result.keys()
        
        return [dict(zip(columns, row)) for row in rows]
    
    def _write_csv(self, cases: List[Dict], output_path: str) -> int:
        """Write cases to CSV format."""
        if not cases:
            logger.warning("No cases to export")
            return 0
        
        # Flatten structure for CSV
        rows = []
        for case in cases:
            row = self._flatten_case(case)
            rows.append(row)
        
        # Get all unique columns
        all_columns = set()
        for row in rows:
            all_columns.update(row.keys())
        
        # Order columns logically
        column_order = [
            'case_id', 'docket_number', 'title', 'court_level', 'court', 'county',
            'district', 'trial_judge', 'source_docket_number',
            'appeal_outcome', 'case_winner_legal', 'case_winner_personal',
            'case_type', 'decision_year', 'decision_month', 'publication_status',
            'issue_count', 'party_count', 'judge_count', 'citation_count', 'statute_count', 'rcw_count',
            'parties_list', 'judges_list',
            'issue_1_summary', 'issue_1_outcome', 'issue_1_winner', 'issue_1_category', 'issue_1_rcw',
            'issue_2_summary', 'issue_2_outcome', 'issue_2_winner', 'issue_2_category', 'issue_2_rcw',
            'issue_3_summary', 'issue_3_outcome', 'issue_3_winner', 'issue_3_category', 'issue_3_rcw',
            'issue_4_summary', 'issue_4_outcome', 'issue_4_winner', 'issue_4_category', 'issue_4_rcw',
            'issue_5_summary', 'issue_5_outcome', 'issue_5_winner', 'issue_5_category', 'issue_5_rcw',
            'source_file', 'extraction_timestamp'
        ]
        
        # Add any columns not in the predefined order
        ordered_columns = [c for c in column_order if c in all_columns]
        remaining = sorted(all_columns - set(ordered_columns))
        ordered_columns.extend(remaining)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=ordered_columns, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)
        
        logger.info(f"Exported {len(rows)} cases to {output_path}")
        return len(rows)
    
    def _write_excel(self, cases: List[Dict], output_path: str) -> int:
        """Write cases to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
        
        if not cases:
            logger.warning("No cases to export")
            return 0
        
        # Flatten cases
        rows = [self._flatten_case(case) for case in cases]
        
        # Get all unique columns with ordering
        column_order = [
            'case_id', 'docket_number', 'title', 'court_level', 'court', 'county',
            'appeal_outcome', 'case_winner_legal', 'case_type', 'issue_count',
            'issue_1_summary', 'issue_1_outcome', 'issue_1_winner', 'issue_1_category',
            'issue_2_summary', 'issue_2_outcome', 'issue_2_winner', 'issue_2_category',
            'issue_3_summary', 'issue_3_outcome', 'issue_3_winner', 'issue_3_category',
            'parties_list', 'judges_list', 'source_file'
        ]
        
        all_columns = set()
        for row in rows:
            all_columns.update(row.keys())
        
        ordered_columns = [c for c in column_order if c in all_columns]
        remaining = sorted(all_columns - set(ordered_columns))
        ordered_columns.extend(remaining)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QA Review"
        
        # Header row with formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(ordered_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(ordered_columns, 1):
                value = row.get(col_name, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths (approximate)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Exported {len(rows)} cases to {output_path}")
        return len(rows)
    
    def _flatten_case(self, case: Dict) -> Dict:
        """Flatten a case record for tabular export."""
        flat = {}
        
        # Copy simple fields
        simple_fields = [
            'case_id', 'docket_number', 'title', 'court_level', 'court', 'county',
            'district', 'source_docket_number',
            'appeal_outcome', 'case_winner_legal', 'case_winner_personal',
            'case_type', 'decision_year', 'decision_month', 'publication_status',
            'issue_count', 'citation_count', 'statute_count', 'rcw_count',
            'source_file', 'extraction_timestamp'
        ]
        
        for field in simple_fields:
            flat[field] = case.get(field)
        
        # Flatten parties
        parties = case.get('parties') or []
        if isinstance(parties, str):
            parties = json.loads(parties) if parties else []
        flat['party_count'] = len(parties)
        flat['parties_list'] = '; '.join([
            f"{p.get('name', '')} ({p.get('role', '')})" 
            for p in parties
        ]) if parties else ''
        
        # Flatten judges
        judges = case.get('judges') or []
        if isinstance(judges, str):
            judges = json.loads(judges) if judges else []
        flat['judge_count'] = len(judges)
        flat['judges_list'] = '; '.join([
            f"{j.get('name', '')} ({j.get('role', '')})" 
            for j in judges
        ]) if judges else ''
        
        # Flatten issues (up to 5)
        issues = case.get('issues') or []
        if isinstance(issues, str):
            issues = json.loads(issues) if issues else []
        
        for i, issue in enumerate(issues[:5], 1):
            prefix = f'issue_{i}_'
            flat[f'{prefix}summary'] = issue.get('summary', '')[:200]
            flat[f'{prefix}outcome'] = issue.get('outcome', '')
            flat[f'{prefix}winner'] = issue.get('winner_legal', '')
            flat[f'{prefix}category'] = f"{issue.get('category', '')} > {issue.get('subcategory', '')}"
            flat[f'{prefix}rcw'] = issue.get('rcw_reference', '')
            flat[f'{prefix}decision'] = issue.get('decision_summary', '')[:200] if issue.get('decision_summary') else ''
        
        return flat
    
    def generate_qa_report(self, output_path: str, limit: int = 100) -> Dict[str, Any]:
        """
        Generate a QA summary report with data quality metrics.
        
        Returns:
            Dictionary with quality metrics
        """
        with self.engine.connect() as conn:
            # Total cases
            total = conn.execute(text("SELECT COUNT(*) FROM cases")).scalar()
            
            # Cases with issues
            with_issues = conn.execute(text(
                "SELECT COUNT(*) FROM cases WHERE issue_count > 0"
            )).scalar()
            
            # Issue count distribution
            issue_dist = conn.execute(text("""
                SELECT issue_count, COUNT(*) as cnt
                FROM cases
                GROUP BY issue_count
                ORDER BY issue_count
            """)).fetchall()
            
            # Outcome distribution
            outcome_dist = conn.execute(text("""
                SELECT issue_outcome, COUNT(*) as cnt
                FROM issues_decisions
                GROUP BY issue_outcome
                ORDER BY cnt DESC
            """)).fetchall()
            
            # Winner role distribution
            winner_dist = conn.execute(text("""
                SELECT winner_legal_role, COUNT(*) as cnt
                FROM issues_decisions
                WHERE winner_legal_role IS NOT NULL
                GROUP BY winner_legal_role
                ORDER BY cnt DESC
            """)).fetchall()
            
            # Cases with potential quality issues
            single_issue = conn.execute(text(
                "SELECT COUNT(*) FROM cases WHERE issue_count = 1"
            )).scalar()
            
            no_issues = conn.execute(text(
                "SELECT COUNT(*) FROM cases WHERE issue_count = 0"
            )).scalar()
        
        report = {
            'generated_at': datetime.now().isoformat(),
            'total_cases': total,
            'cases_with_issues': with_issues,
            'cases_without_issues': no_issues,
            'single_issue_cases': single_issue,
            'issue_count_distribution': {str(r[0]): r[1] for r in issue_dist},
            'outcome_distribution': {str(r[0]): r[1] for r in outcome_dist},
            'winner_role_distribution': {str(r[0]): r[1] for r in winner_dist},
            'quality_flags': {
                'pct_single_issue': round(single_issue / total * 100, 1) if total else 0,
                'pct_no_issues': round(no_issues / total * 100, 1) if total else 0,
            }
        }
        
        # Write report
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2)
        
        logger.info(f"QA report written to {output_path}")
        return report


def main():
    parser = argparse.ArgumentParser(
        description='Quality Assurance Export for Extraction Verification'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        default='qa_review.csv',
        help='Output file path (default: qa_review.csv)'
    )
    
    parser.add_argument(
        '--format', '-f',
        type=str,
        choices=['csv', 'excel'],
        default='csv',
        help='Output format (default: csv)'
    )
    
    parser.add_argument(
        '--limit', '-l',
        type=int,
        default=None,
        help='Maximum number of cases to export'
    )
    
    parser.add_argument(
        '--case-ids',
        type=str,
        default=None,
        help='Comma-separated list of specific case IDs to export'
    )
    
    parser.add_argument(
        '--report',
        action='store_true',
        help='Generate QA summary report instead of case export'
    )
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    exporter = QAExporter()
    
    if args.report:
        report_path = args.output.replace('.csv', '_report.json').replace('.xlsx', '_report.json')
        report = exporter.generate_qa_report(report_path)
        print(f"\nQA Report Summary:")
        print(f"  Total cases: {report['total_cases']}")
        print(f"  Cases with issues: {report['cases_with_issues']}")
        print(f"  Single-issue cases: {report['single_issue_cases']} ({report['quality_flags']['pct_single_issue']}%)")
        print(f"  Cases without issues: {report['cases_without_issues']} ({report['quality_flags']['pct_no_issues']}%)")
        print(f"\n  Issue outcome distribution:")
        for outcome, count in report['outcome_distribution'].items():
            print(f"    {outcome}: {count}")
    else:
        case_ids = None
        if args.case_ids:
            case_ids = [int(x.strip()) for x in args.case_ids.split(',')]
        
        count = exporter.export_cases(
            output_path=args.output,
            case_ids=case_ids,
            limit=args.limit,
            format=args.format
        )
        print(f"\nExported {count} cases to {args.output}")


if __name__ == '__main__':
    main()
